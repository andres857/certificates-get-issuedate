from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import logging
from datetime import datetime

def create_excel_template(current_path: str, name_report: str = "report") -> Path:
    """
    Crea un archivo Excel con múltiples hojas para reportes completos
    """
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s')
    
    try:
        excel_path = Path(current_path) / f'{name_report}.xlsx'
        
        if excel_path.exists():
            logging.info(f"El archivo {excel_path} ya existe. No se creará uno nuevo.")
            return excel_path
        
        wb = Workbook()
        
        # Hoja 1: Resumen general
        ws_summary = wb.active
        ws_summary.title = "Resumen"
        create_summary_sheet(ws_summary)
        
        # Hoja 2: Certificados procesados exitosamente
        ws_processed = wb.create_sheet("Procesados")
        create_processed_sheet(ws_processed)
        
        # Hoja 3: Duplicados encontrados
        ws_duplicates = wb.create_sheet("Duplicados")
        create_duplicates_sheet(ws_duplicates)
        
        # Hoja 4: Errores encontrados
        ws_errors = wb.create_sheet("Errores")
        create_errors_sheet(ws_errors)
        
        wb.save(str(excel_path))
        logging.info(f"Archivo Excel creado exitosamente en: {excel_path}")
        
        return excel_path
    
    except Exception as e:
        logging.error(f"Error al crear el archivo Excel: {str(e)}")
        raise

def create_summary_sheet(ws):
    """Crea la hoja de resumen con estadísticas generales"""
    # Título principal
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "REPORTE DE PROCESAMIENTO DE CERTIFICADOS"
    title_cell.font = Font(bold=True, size=16, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Información de fecha
    ws.merge_cells('A2:F2')
    date_cell = ws['A2']
    date_cell.value = f"Generado el: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    date_cell.font = Font(italic=True)
    date_cell.alignment = Alignment(horizontal="center")
    
    # Headers de estadísticas
    stats_headers = [
        "Estadística", "Cantidad", "Porcentaje", "Observaciones"
    ]
    
    for col, header in enumerate(stats_headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Filas de estadísticas (se llenarán dinámicamente)
    stats_rows = [
        "Total de archivos encontrados",
        "Certificados procesados exitosamente", 
        "Duplicados detectados",
        "Errores de procesamiento",
        "Archivos ya procesados (omitidos)"
    ]
    
    for idx, stat in enumerate(stats_rows, 5):
        ws.cell(row=idx, column=1, value=stat)
        ws.cell(row=idx, column=2, value=0)  # Se actualizará dinámicamente
        ws.cell(row=idx, column=3, value="0%")  # Se actualizará dinámicamente
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 30

def create_processed_sheet(ws):
    """Crea la hoja de certificados procesados exitosamente"""
    headers = [
        "Archivo Original", "Archivo Renombrado", "Identificación", 
        "Nombre", "Fecha Emisión", "Fecha Expiración", "Timestamp"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Ajustar ancho de columnas
    column_widths = [30, 40, 15, 25, 15, 15, 20]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

def create_duplicates_sheet(ws):
    """Crea la hoja de duplicados detectados"""
    headers = [
        "Archivo Duplicado", "Archivo Original", "Identificación", 
        "Nombre", "Fecha Emisión", "Fecha Expiración", "Razón", "Timestamp"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Ajustar ancho de columnas
    column_widths = [30, 30, 15, 25, 15, 15, 30, 20]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

def create_errors_sheet(ws):
    """Crea la hoja de errores"""
    headers = [
        "Archivo", "Tipo de Error", "Descripción del Error", 
        "Identificación", "Datos Parciales", "Timestamp"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Ajustar ancho de columnas
    column_widths = [30, 20, 40, 15, 30, 20]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

def insert_processed_certificate(excel_path: Path, data: dict, original_name: str, renamed_file: str) -> None:
    """Inserta un certificado procesado exitosamente"""
    try:
        wb = load_workbook(str(excel_path))
        ws = wb["Procesados"]
        
        last_row = ws.max_row + 1
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        ws.cell(row=last_row, column=1, value=original_name)
        ws.cell(row=last_row, column=2, value=renamed_file)
        ws.cell(row=last_row, column=3, value=data.get('identification', ''))
        ws.cell(row=last_row, column=4, value=data.get('name', ''))
        ws.cell(row=last_row, column=5, value=data.get('issue_date', ''))
        ws.cell(row=last_row, column=6, value=data.get('expiration_date', ''))
        ws.cell(row=last_row, column=7, value=timestamp)
        
        wb.save(str(excel_path))
        
    except Exception as e:
        logging.error(f"Error al insertar certificado procesado: {str(e)}")

def insert_duplicate_certificate(excel_path: Path, data: dict, duplicate_file: str, 
                                original_file: str, reason: str) -> None:
    """Inserta un duplicado detectado"""
    try:
        wb = load_workbook(str(excel_path))
        ws = wb["Duplicados"]
        
        last_row = ws.max_row + 1
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        ws.cell(row=last_row, column=1, value=duplicate_file)
        ws.cell(row=last_row, column=2, value=original_file)
        ws.cell(row=last_row, column=3, value=data.get('identification', ''))
        ws.cell(row=last_row, column=4, value=data.get('name', ''))
        ws.cell(row=last_row, column=5, value=data.get('issue_date', ''))
        ws.cell(row=last_row, column=6, value=data.get('expiration_date', ''))
        ws.cell(row=last_row, column=7, value=reason)
        ws.cell(row=last_row, column=8, value=timestamp)
        
        wb.save(str(excel_path))
        
    except Exception as e:
        logging.error(f"Error al insertar duplicado: {str(e)}")

def insert_error_certificate(excel_path: Path, filename: str, error_type: str, 
                           error_description: str, partial_data: dict = None) -> None:
    """Inserta un error de procesamiento"""
    try:
        wb = load_workbook(str(excel_path))
        ws = wb["Errores"]
        
        last_row = ws.max_row + 1
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        ws.cell(row=last_row, column=1, value=filename)
        ws.cell(row=last_row, column=2, value=error_type)
        ws.cell(row=last_row, column=3, value=error_description)
        ws.cell(row=last_row, column=4, value=partial_data.get('identification', '') if partial_data else '')
        ws.cell(row=last_row, column=5, value=str(partial_data) if partial_data else '')
        ws.cell(row=last_row, column=6, value=timestamp)
        
        wb.save(str(excel_path))
        
    except Exception as e:
        logging.error(f"Error al insertar error: {str(e)}")

def update_summary_stats(excel_path: Path, stats: dict) -> None:
    """Actualiza las estadísticas del resumen"""
    try:
        wb = load_workbook(str(excel_path))
        ws = wb["Resumen"]
        
        total = sum(stats.values())
        
        # Mapeo de estadísticas a filas
        stat_mapping = {
            'total_files': 5,
            'processed': 6, 
            'duplicates': 7,
            'errors': 8,
            'already_processed': 9
        }
        
        for stat_key, row in stat_mapping.items():
            if stat_key in stats:
                count = stats[stat_key]
                percentage = (count / total * 100) if total > 0 else 0
                
                ws.cell(row=row, column=2, value=count)
                ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
        
        wb.save(str(excel_path))
        
    except Exception as e:
        logging.error(f"Error al actualizar estadísticas: {str(e)}")

# Mantener compatibilidad con código existente
def insert_certificate_data(excel_path: Path, data: dict) -> None:
    """
    Función de compatibilidad - redirige a insert_error_certificate
    """
    error_msg = data.get('message_error', data.get('error_message', 'Error desconocido'))
    insert_error_certificate(
        excel_path, 
        data.get('name', 'Archivo desconocido'),
        'Error de procesamiento',
        error_msg,
        data
    )